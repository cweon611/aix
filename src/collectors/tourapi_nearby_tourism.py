"""전남·광주 관광지 풀을 만들고 인기명소에 인기도를 얹는다.

식당 상세의 '주변 관광 정보'가 쓸 데이터다. 예전에는 두 방식을 거쳤다.
  1) 거리 5km 반경 조회 → 소소한 곳만, 명소는 빠짐
  2) 인기명소 목록만 지오코딩 → 유명하지만 31곳뿐이라 완도·진도·강진 등
     인기명소가 없는 지역의 식당은 주변에 아무것도 안 뜸(266곳 중 95곳이 0~1곳)

그래서 둘을 합친다. TourAPI로 전남·광주의 관광지(12)·문화시설(14)을 폭넓게
받아 풀을 채우고(어느 식당에서도 가까운 곳이 잡히도록), 그 위에 인기명소
목록의 인기도를 이름으로 맞춰 얹는다. 화면은 이 풀에서 지점마다 가까운 순으로
자른다. 인기도가 있는 곳은 배지로 앞세운다.

거리·식당은 수백 곳이지만 관광 풀은 한 번 받아 굳히면 되고, 웹 런타임에서는
좌표로 거리만 재므로 키·서버가 필요 없다.

입력 : 카카오톡 받은 파일/전남인기관광명소.xlsx (인기도 오버레이, 없어도 됨)
출력 : data/processed/nearby_tourism.json  (관광지 풀. 웹 빌드가 tourism.json으로 굳힌다)

실행: python -m src.collectors.tourapi_nearby_tourism
"""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path

from src.collectors.tourapi_region_food import (
    TourApiError,
    fetch_area_list,
    fetch_region_codes,
)
from src.config import DATA_PROCESSED_DIR

OUTPUT = DATA_PROCESSED_DIR / "nearby_tourism.json"
DEFAULT_XLSX = (
    Path.home() / "OneDrive" / "문서" / "카카오톡 받은 파일" / "전남인기관광명소.xlsx"
)

CONTENT_TYPES = {12: "관광지", 14: "문화시설"}


def load_popularity(path: Path) -> dict[str, tuple[float, str]]:
    """이름 -> (인기 비율, 구분). 이름 맞춤용으로 공백·괄호를 지운 키도 만든다."""
    if not path.exists():
        print(f"  인기명소 파일 없음({path.name}). 인기도 없이 풀만 만듭니다.")
        return {}
    from openpyxl import load_workbook

    ws = load_workbook(path, data_only=True).active
    out: dict[str, tuple[float, str]] = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r[1]:
            continue
        name = _norm(str(r[1]))
        try:
            ratio = float(r[4]) if r[4] is not None else 0.0
        except (TypeError, ValueError):
            ratio = 0.0
        out[name] = (ratio, str(r[2]).strip() if r[2] else "관광지")
    return out


def _norm(s: str) -> str:
    return re.sub(r"[\s()（）]", "", s)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    args = parser.parse_args()

    popularity = load_popularity(args.xlsx)
    regions = fetch_region_codes()
    print("지역코드:", ", ".join(f"{r['name']}({r['code']})" for r in regions))

    pool: dict[str, dict] = {}
    for region in regions:
        for type_id, type_label in CONTENT_TYPES.items():
            try:
                rows = fetch_area_list(region["code"], type_id)
            except TourApiError as exc:
                print(f"  {region['name']} {type_label} 실패: {exc}")
                continue
            print(f"  {region['name']} {type_label}: {len(rows)}건")
            for r in rows:
                cid = r.get("contentid")
                if not cid or cid in pool:
                    continue
                if not r.get("mapx") or not r.get("mapy"):
                    continue
                name = (r.get("title") or "").strip()
                pop, pop_type = popularity.get(_norm(name), (0.0, ""))
                image = (r.get("firstimage") or "").strip()
                if image.startswith("http://"):
                    image = "https://" + image[len("http://"):]
                pool[cid] = {
                    "id": cid,
                    "title": name,
                    "type": pop_type or type_label,
                    "popularity": pop,
                    "addr": (r.get("addr1") or "").strip(),
                    "lat": float(r["mapy"]),
                    "lon": float(r["mapx"]),
                    "image": image,
                    "tel": (r.get("tel") or "").strip(),
                }

    spots = sorted(pool.values(), key=lambda x: -x["popularity"])
    OUTPUT.write_text(json.dumps(spots, ensure_ascii=False, indent=2), encoding="utf-8")

    tagged = sum(1 for s in spots if s["popularity"] > 0)
    imaged = sum(1 for s in spots if s["image"])
    matched = len({_norm(s["title"]) for s in spots} & set(popularity))
    print(f"\n관광지 풀 {len(spots)}곳 저장 -> {OUTPUT}")
    print(f"  인기도 붙은 곳 {tagged}곳 / 인기명소 {len(popularity)}곳 중 {matched}곳 매칭")
    print(f"  사진 있는 곳 {imaged}곳")


if __name__ == "__main__":
    main()
